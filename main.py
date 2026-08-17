# -*- coding: utf-8 -*-
"""
ecommerce-file-reader 插件
==========================

读取工作区中的电商文件，并返回结构化内容。

支持的文件类型:
  - 图片 (jpg / jpeg / png / webp): 文件名称、图片尺寸、图片格式
  - Excel (xlsx / xls):             表格字段、数据行数、前 20 行数据
  - CSV  (csv):                     字段、数据内容
  - PDF  (pdf):                     页数、文本内容
  - 文本 (txt / md):                文件文本内容

工具:
  read_file(file_path)  -> 读取指定文件路径并返回结构化内容

直接运行:
  python main.py <文件路径>
"""

import csv
import json
import sys
from pathlib import Path

# 支持的文件扩展名
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}
EXCEL_EXTS = {".xlsx", ".xls"}
CSV_EXTS = {".csv"}
PDF_EXTS = {".pdf"}
TEXT_EXTS = {".txt", ".md"}

# 输出限制
MAX_TEXT_CHARS = 100_000    # 文本 / PDF 内容最多返回字符数
MAX_CSV_ROWS = 100          # CSV 最多返回的数据行数
EXCEL_HEAD_ROWS = 20        # Excel 返回的前 N 行数据

TOOL_NAMES = ["read_file"]


def _sanitize(value):
    """把单元格值转换为可 JSON 序列化的 Python 对象。"""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if hasattr(value, "isoformat"):  # datetime / date / time
        return value.isoformat()
    if isinstance(value, (list, tuple)):
        return [_sanitize(v) for v in value]
    return str(value)


def _detect_encoding(path):
    """按优先级探测文本编码（UTF-8 -> GB18030 -> Latin-1）。"""
    for enc in ("utf-8-sig", "utf-8", "gb18030", "gbk", "latin-1"):
        try:
            with open(path, "rb") as f:
                f.read(64 * 1024).decode(enc)
            return enc
        except (UnicodeDecodeError, OSError):
            continue
    return "latin-1"


# ---------------------------------------------------------------------------
# 1. 图片: jpg / jpeg / png / webp
# ---------------------------------------------------------------------------
def _read_image(path):
    """读取图片: 文件名称（外层）、尺寸、格式。"""
    try:
        from PIL import Image
    except ImportError as e:
        raise RuntimeError("缺少依赖 Pillow，请先执行: pip install -r requirements.txt") from e

    with Image.open(path) as img:
        width, height = img.size
        return {
            "width": width,
            "height": height,
            "dimensions": "{0} x {1}".format(width, height),
            "format": img.format,
        }


# ---------------------------------------------------------------------------
# 2. Excel: xlsx / xls
# ---------------------------------------------------------------------------
def _build_excel_sheet(sheet_name, rows):
    """从原始行数据中提取表头字段、总数据行数和前 N 行数据。"""
    header_idx = None
    for i, row in enumerate(rows):
        if any(v is not None and str(v).strip() != "" for v in row):
            header_idx = i
            break

    if header_idx is None:
        return {"sheet_name": sheet_name, "fields": [], "total_rows": 0, "first_20_rows": []}

    fields = [str(v) if v is not None else "" for v in rows[header_idx]]
    data = [
        row
        for row in rows[header_idx + 1:]
        if any(v is not None and str(v).strip() != "" for v in row)
    ]
    return {
        "sheet_name": sheet_name,
        "fields": fields,
        "total_rows": len(data),
        "first_20_rows": data[:EXCEL_HEAD_ROWS],
    }


def _read_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("缺少依赖 openpyxl，请先执行: pip install -r requirements.txt") from e

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                if row is None:
                    continue
                rows.append([_sanitize(v) for v in row])
            sheets.append(_build_excel_sheet(ws.title, rows))
        return {"sheets": sheets}
    finally:
        wb.close()


def _read_xls(path):
    try:
        import xlrd
    except ImportError as e:
        raise RuntimeError("缺少依赖 xlrd，请先执行: pip install -r requirements.txt") from e

    book = xlrd.open_workbook(str(path))
    sheets = []
    for sheet in book.sheets():
        rows = []
        for r in range(sheet.nrows):
            row = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        row.append(xlrd.xldate_as_datetime(cell.value, book.datemode).isoformat(" "))
                    except Exception:
                        row.append(str(cell.value))
                elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    row.append(bool(cell.value))
                elif cell.ctype == xlrd.XL_CELL_ERROR:
                    row.append("#ERR({0})".format(cell.value))
                else:
                    row.append(_sanitize(cell.value))
            rows.append(row)
        sheets.append(_build_excel_sheet(sheet.name, rows))
    return {"sheets": sheets}


def _read_excel(path):
    return _read_xlsx(path) if path.suffix.lower() == ".xlsx" else _read_xls(path)


# ---------------------------------------------------------------------------
# 3. CSV
# ---------------------------------------------------------------------------
def _read_csv(path):
    """读取 CSV: 字段 + 数据内容（自动识别编码与分隔符）。"""
    encoding = _detect_encoding(path)
    with open(path, "r", encoding=encoding, errors="replace", newline="") as f:
        sample = f.read(8192)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        rows = [row for row in reader if any(cell.strip() for cell in row)]

    fields = rows[0] if rows else []
    data = rows[1:]
    return {
        "fields": fields,
        "total_rows": len(data),
        "rows": data[:MAX_CSV_ROWS],
        "encoding": encoding,
    }


# ---------------------------------------------------------------------------
# 4. PDF
# ---------------------------------------------------------------------------
def _read_pdf(path):
    """读取 PDF: 页数 + 每页文本内容。"""
    try:
        from pypdf import PdfReader
    except ImportError:
        try:
            from PyPDF2 import PdfReader
        except ImportError as e:
            raise RuntimeError("缺少依赖 pypdf，请先执行: pip install -r requirements.txt") from e

    reader = PdfReader(str(path))
    page_count = len(reader.pages)
    pages = []
    total_chars = 0
    truncated = False
    for i, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""
        pages.append({"page": i, "text": text})
        total_chars += len(text)
        if total_chars >= MAX_TEXT_CHARS:
            truncated = True
            break
    return {"page_count": page_count, "pages": pages, "truncated": truncated}


# ---------------------------------------------------------------------------
# 5. TXT / MD
# ---------------------------------------------------------------------------
def _read_text(path):
    """读取纯文本: 文件文本内容（自动识别编码）。"""
    encoding = _detect_encoding(path)
    content = path.read_text(encoding=encoding, errors="replace")
    truncated = False
    if len(content) > MAX_TEXT_CHARS:
        content = content[:MAX_TEXT_CHARS]
        truncated = True
    return {"content": content, "encoding": encoding, "truncated": truncated}


# ---------------------------------------------------------------------------
# 工具入口
# ---------------------------------------------------------------------------
def read_file(file_path):
    """读取指定文件路径并返回结构化内容。

    参数:
        file_path (str): 要读取的文件路径（相对或绝对路径）

    返回:
        dict: 结构化读取结果，格式如下
            {
              "success": True,
              "file_name": "...",
              "file_type": "xlsx",
              "file_size": 12345,
              ... 按文件类型附加的数据 ...
            }
    """
    path = Path(file_path).expanduser()
    if not path.exists():
        return {"success": False, "error": "文件不存在: {0}".format(file_path)}
    if not path.is_file():
        return {"success": False, "error": "路径不是文件: {0}".format(file_path)}

    ext = path.suffix.lower()
    base = {
        "success": True,
        "file_name": path.name,
        "file_type": ext.lstrip("."),
        "file_size": path.stat().st_size,
    }

    try:
        if ext in IMAGE_EXTS:
            data = _read_image(path)
        elif ext in EXCEL_EXTS:
            data = _read_excel(path)
        elif ext in CSV_EXTS:
            data = _read_csv(path)
        elif ext in PDF_EXTS:
            data = _read_pdf(path)
        elif ext in TEXT_EXTS:
            data = _read_text(path)
        else:
            return {**base, "success": False, "error": "不支持的文件类型: {0}".format(ext)}
    except Exception as e:
        return {**base, "success": False, "error": "读取失败: {0}".format(e)}

    return {**base, **data}


def execute(tool_name, **kwargs):
    """按工具名分发调用（供宿主平台按工具名调用）。"""
    if tool_name == "read_file":
        return read_file(kwargs.get("file_path"))
    return {"success": False, "error": "未知工具: {0}".format(tool_name)}


def main():
    if len(sys.argv) < 2:
        print("用法: python main.py <文件路径>")
        sys.exit(1)
    result = read_file(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
